from selenium import webdriver
from selenium.webdriver.common.by import By
from twocaptcha import TwoCaptcha
import time
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import json
import random
from openpyxl import load_workbook,Workbook
import requests
from groq import Groq
import traceback
from pathlib import Path

# Chech internet
def check_internet():
    test_sites = [
        "https://1.1.1.1",  # Cloudflare DNS (very reliable)
        "https://www.cloudflare.com",  # Cloudflare official site
        "https://example.com",  # Example website (simple and lightweight)
        "https://www.bing.com"  # Bing as an alternative to Google
    ]
    
    for site in test_sites:
        try:
            response = requests.get(site, timeout=10)
            if response.status_code == 200:
                return True
            else:
                print(f"Failed to connect to {site}. HTTP Status Code: {response.status_code}")
        except requests.RequestException as e:
           pass
    
    # print(f"Error connecting  : ")
    
    return False
# intialize driver
def initialize_browser(headless=False,browsersnum=0):   
    print(f"\n[+] Initializing driver")

    options = Options()
    service = Service(ChromeDriverManager().install())

    if headless:
        options.add_argument("--headless")
    
    options.add_argument("window-size=1920x1080")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--lang=en-US")
    options.set_capability("goog:loggingPrefs", {"browser": "INFO"})
    agentbrowsers=[
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.5481.77 Safari/537.36",
        "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1 Safari/605.1.15",
        "user-agent=Mozilla/5.0 (Linux; Android 12; SM-G996U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Mobile Safari/537.36"
    ]
    options.add_argument(agentbrowsers[browsersnum])
    driver = webdriver.Chrome(options=options, service=service)
    return driver

# CLOUDFLARE BYPASS LOGIC
api_key = "31300b1641332536703a487a87b3f4f7"  
solver = TwoCaptcha(api_key)
def solve_captcha(site_key, url):
    """
    Solves CAPTCHA using 2Captcha Python SDK.
    Args:
        site_key (str): CAPTCHA site key.
        url (str): URL where the CAPTCHA is located.
    Returns:
        str: CAPTCHA token if successful.
    """
    try:
        print("Attempting to solve CAPTCHA...")
        result = solver.turnstile(sitekey=site_key, url=url)
        print("CAPTCHA solved successfully.")
        return result["code"]
    except Exception as e:
        print(f"CAPTCHA solving failed: {e}")
        return None

def bypass_cloudflare():
    print("[+] Attempting Cloudflare Bypass")

    params = get_captcha_params()
    if params:
        token = solve_captcha(params)

        if token:
            send_token(token)
            time.sleep(5)
            print("[+] Bypassed Successfully!")
        else:
            print("[-] Failed to solve captcha")
    else:
        print("[-] Failed to intercept parameters")

def get_captcha_params():
    intercept_script = """
        console.clear = () => console.log('Console was cleared')
        const i = setInterval(()=>{
        if (window.turnstile)
            console.log('success!!')
            {clearInterval(i)
                window.turnstile.render = (a,b) => {
                let params = {
                    sitekey: b.sitekey,
                    pageurl: window.location.href,
                    data: b.cData,
                    pagedata: b.chlPageData,
                    action: b.action,
                    userAgent: navigator.userAgent,
                }
                console.log('intercepted-params:' + JSON.stringify(params))
                window.cfCallback = b.callback
                return        }
        }
    },50)
    """
    params = None

    while not params:
        print("[+] Attempting to Get params")
        driver.refresh()
        driver.execute_script(intercept_script)
        time.sleep(5)

        logs =driver.get_log("browser")
        for log in logs:
            if "intercepted-params:" in log["message"]:
                log_entry = log["message"].encode("utf-8").decode("unicode_escape")
                match = re.search(r"intercepted-params:({.*?})", log_entry)
                if match:
                    json_string = match.group(1)
                    params = json.loads(json_string)
                    break

    print("[+] Parameters received")
    return params

def solve_captcha(params):
    solver = TwoCaptcha(api_key)
    try:
        result = solver.turnstile(
            sitekey=params["sitekey"],
            url=params["pageurl"],
            action=params["action"],
            data=params["data"],
            pagedata=params["pagedata"],
            useragent=params["userAgent"],
        )
        print(f"[+] Captcha solved")
        return result["code"]
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def send_token(token):
    script = f"cfCallback('{token}')"
    driver.execute_script(script)

# work like main function
def extracting_jobs(about_me,job_urls_list,ignore_companies_list,jobs_per_company,max_items):
    """ Extracting URLs from the provided urls in links list."""
    # intialization workbook for saving extracted data
    try:
        # Get the directory where scrapper.py is located
        current_dir = Path(__file__).parent  # Points to scrapapply\indeed
        EXCEL_PATH = current_dir / "static" / "indeed" / "download" / "indeed_jobs.xlsx"
        wb = Workbook()
        sheet = wb.active
        header = ["Company", "Position", "Location", "Link", "Match %"]
        sheet.append(header)
        wb.save(EXCEL_PATH)
        print("[+] Workbook initialized successfully")
    except Exception as e:
        print(f"[-] Error initializing workbook: {e}")
    
    # var for checking max times this will be increase
    saved_jobs=0
    try:
        # Initialize Selenium WebDriver
        global driver
        driver = initialize_browser()
        driver.implicitly_wait(4)
        for link in job_urls_list:
            try:
                if saved_jobs >= max_items:
                    break
                # Wait until the internet is stable              
                while not check_internet():
                    time.sleep(10)  
                    driver.refresh()                
               
                driver.get(link)
                time.sleep(random.uniform(2, 5))  # Randomized delay
                # CAPTCHA handling
                if "Additional Verification Required" in driver.page_source:
                    time.sleep(5)
                    try:
                        bypass_cloudflare()
                    except Exception as captcha_error:
                        print(f"CAPTCHA handling error for {link}: {captcha_error}")  
                        continue      
                crawler(about_me,ignore_companies_list,jobs_per_company,max_items,saved_jobs,EXCEL_PATH)
            except Exception as e:
                print(f"Error processing {link}: and the current driver urls are : {driver.current_url}")

    except Exception as e:
        print(f"Unexpected error: {e}")

# this function are crawler page one by one and scrip jobs..
def crawler(about_me, ignore_companies_list, jobs_per_company, max_items, saved_jobs, EXCEL_PATH):    
    list_of_company = []
    list_of_link = []
    list_of_title = []
    list_of_location = []
    count_company_jobs = []

    ignore_companies = ignore_companies_list
    unique_ids = set()

    page_number = 1
    while True: 
        if saved_jobs >= max_items:
            break

        if "Additional Verification Required" in driver.page_source:
            time.sleep(5)
            try:
                bypass_cloudflare()
            except Exception as captcha_error:
                print(f"CAPTCHA handling error: {captcha_error}")
                continue
        
        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "body"))
        )
        
        while not check_internet():
            time.sleep(10)
            driver.refresh()

        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(2,3))

            company_elems = driver.find_elements(By.XPATH, "//span[@data-testid='company-name']")
            link_elems = driver.find_elements(By.XPATH, "//a[contains(@class,'jcs-JobTitle')]")
            title_elems = driver.find_elements(By.XPATH, "//h2[contains(@class, 'jobTitle')]")
            location_elems = driver.find_elements(By.XPATH, "//div[@data-testid='text-location']")

            if len(company_elems) != len(link_elems) or len(title_elems) != len(location_elems):
                print("Length of all elements are not the same! Skipping page.")
                continue
        except Exception as e:
            print(f"Error: Page may not have loaded properly or elements do not match. {e}")
            break

        for company_elem, link_elem, title_elem, location_elem in zip(company_elems, link_elems, title_elems, location_elems):
            href = link_elem.get_attribute("href")     
            try:
                # Extract job ID
                if "/view/" in href:
                    job_id = href.split("/view/")[1].split("/")[0]
                elif "=" in href:
                    job_id = href.split("=")[1].split("&")[0]
                else:
                    continue
                
                if saved_jobs >= max_items:
                    break

                company = company_elem.text.strip()
                title = title_elem.text.strip()
                location = location_elem.text.strip()

                count = count_company_jobs.count(company)
                
                if count >= jobs_per_company or job_id in unique_ids or company in ignore_companies:
                    continue

                # Only after all checks passed
                unique_ids.add(job_id)
                count_company_jobs.append(company)

                list_of_company.append(company)
                list_of_link.append(href)
                list_of_title.append(title)
                list_of_location.append(location)

                saved_jobs += 1

                print(f"len of titles : {len(list_of_company)}")
                if len(list_of_company) >= 10:
                    print(f"\nWait for deepseek response it take at least 3 or 4s...\n")
                    process_jobs(list_of_company, list_of_link, list_of_title, list_of_location, about_me, EXCEL_PATH)
                    list_of_company.clear()
                    list_of_link.clear()
                    list_of_title.clear()
                    list_of_location.clear()
            except Exception as e:
                print(f"Error processing job link: {e}")

        time.sleep(random.uniform(2, 5))
        page_number += 1

        list_of_xpaths = [
            f"//a[@data-testid='pagination-page-{page_number}']",  
            f"(//a[normalize-space()='{page_number}'])[1]"
        ]

        for xpath in list_of_xpaths:
            try:
                paginator = driver.find_element(By.XPATH, xpath)
                paginator.click()
                print(f"[+] Successfully clicked pagination {page_number}")
                break
            except Exception as e:
                continue
        else:
            print(f"[-] All selectors failed for page {page_number}")
            driver.execute_script("window.scrollTo(0, 0);")
            driver.save_screenshot(f"screenshot/page_{page_number}.png")

            if list_of_company:
                print(f"\nlast page Wait for deepseek response it take at least 3 or 4s...\n")
                process_jobs(list_of_company, list_of_link, list_of_title, list_of_location, about_me,EXCEL_PATH)
                list_of_company.clear()
                list_of_link.clear()
                list_of_title.clear()
                list_of_location.clear()
            break

# function that call ai for checking titles to matching or not
def request_deepseek(prompt):
    client = Groq(api_key='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')
    messages = [{"role": "user", "content": prompt}]
    
    try:
        completion = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=messages,
            temperature=0,
            max_tokens=1024,
            top_p=1,
            stream=False
        )

        # Extract and clean the response
        response_text = completion.choices[0].message.content.strip()
        return response_text

    except Exception as e:
        print("\nError:", e)
        print(traceback.format_exc())
        return None

def process_jobs(list_of_company,list_of_link,list_of_title,list_of_location,about_me,EXCEL_PATH):
    my_req = f"""
Below is the 'About Me'.
For each job title, provide **percentage**:
1. **Match Percentage** – How well this job title matches my profile.
Return only the highest percentage (Match) for each job title,and your response must in a single line of numbers separated by spaces and keep in mind  without any explaination info. For example:
Here are the highest percentages
95 85 30 65...
About Me:
{about_me}

Job Titles:
{"\n".join(list_of_title)}
"""
    # call the deepseek
    res_deepseek = request_deepseek(my_req)
    percentages = re.findall(r'\b\d+\b', res_deepseek)
    
    # load excel for updates
    wb=load_workbook(EXCEL_PATH)
    sheet=wb.active

    # iterate
    for  company,link,title,location,percentage in zip(list_of_company, list_of_link, list_of_title, list_of_location,percentages):
        sheet.append([company,link,title,location,percentage])
        wb.save(EXCEL_PATH)
        print(f"Saved Sucessfully: {title}")
        print(f"_____________________________________________________________________")

    